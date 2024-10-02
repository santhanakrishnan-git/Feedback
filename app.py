from flask import Flask, make_response, render_template, send_file, redirect, url_for, request, flash, session
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from forms import FeedbackForm, LoginForm
from datetime import datetime, timedelta
from models import Feedback
from db import db  # Import db from the new db.py file
from flask_migrate import Migrate
from flask_mail import Mail, Message
from dotenv import load_dotenv
import os

app = Flask(__name__)
#load .env file
load_dotenv()

#Get values from .env file
GUEST_USERNAME = os.getenv('GUEST_USERNAME')
GUEST_PASSWORD= os.getenv("GUEST_PASSWORD")

#print(f"Database URL: {GUEST_USERNAME}")
#print(f"Secret Key: {GUEST_PASSWORD}")

# Configuration for the database and secret key
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///feedback.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your_secret_key'
# Looking to send emails in production? Check out our Email API/SMTP product!
app.config['MAIL_SERVER']='sandbox.smtp.mailtrap.io'
app.config['MAIL_PORT'] = 2525
app.config['MAIL_USERNAME'] = '999a7babd90f63'
app.config['MAIL_PASSWORD'] = '24f83c47385ea6'
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USE_SSL'] = False
app.config['MAIL_DEFAULT_SENDER']='test@example.com'
mail= Mail(app)
# Initialize SQLAlchemy
db.init_app(app)
migrate = Migrate(app, db)

# Set session timeout to 1 minute
app.permanent_session_lifetime = timedelta(minutes=3)

# Ensure the database is created within the app context
with app.app_context():
    db.create_all()

# Home page with link to feedback form
@app.route('/')
def index():
    return render_template('index.html')
# Guest Sign in 
@app.route('/guest_signin', methods=['GET', 'POST'])
def guest_signin():
    if request.method=='POST':
        username=request.form['username']
        password=request.form['password']

        if username== GUEST_USERNAME and password == GUEST_PASSWORD :
            session['logged_in'] = True
            return redirect(url_for('feedback_form')) #redirect to feedback form
        else:
            flash('Invalid Username or Password. Please try again.')
    return render_template('guest_signin.html')

# Route for feedback form
@app.route('/feedback', methods=['GET', 'POST'])
def feedback_form():
    if not session.get('logged_in'):
        return redirect(url_for('guest_signin'))
    else:
        form = FeedbackForm()
        if form.validate_on_submit():
            local_time=request.form.get('localTime')
            feedback = Feedback(
                customer_name=form.customer_name.data,
                project=form.project.data,
                product=form.product.data,
                aipl_job_reference=form.aipl_job_reference.data,
                customer_representative=form.customer_representative.data,
                designation=form.designation.data,
                pre_order_service=form.pre_order_service.data,
                post_order_handling=form.post_order_handling.data,
                on_time_delivery=form.on_time_delivery.data,
                quality_sustenance=form.quality_sustenance.data,
                overall_response=form.overall_response.data,
                final_documentation=form.final_documentation.data,
                appraisal=form.appraisal.data,
                compliance=form.compliance.data,
                comments=form.comments.data,
                timestamp=local_time
            )
            db.session.add(feedback)
            db.session.commit()

            #mapping radio button options
            options_map = {
                5: 'Excellent',
                4: 'Better',
                3: 'Good',
                2: 'Average',
                1: 'Poor'
            }
            #collecting form data
            form_data={
                'Customer Name' : form.customer_name.data,
                'Project' : form.project.data,
                'Product' : form.product.data,
                'AIPL Job Reference' : form.aipl_job_reference.data,
                'Customer Representative' : form.customer_representative.data,
                'Designation' : form.designation.data,
                'Pre-Order Service' : options_map.get(form.pre_order_service.data),
                'Post-Order Handling' : options_map.get(form.post_order_handling.data),
                'On-Time Delivery' : options_map.get(form.on_time_delivery.data),
                'Quality Sustenance' : options_map.get(form.quality_sustenance.data),
                'Overall Response' : options_map.get(form.overall_response.data),
                'Final Documentation' : options_map.get(form.final_documentation.data),
                'Appraisal' : options_map.get(form.appraisal.data),
                'Standard Compliance' : options_map.get(form.compliance.data),
                'Comments' : form.comments.data,
                'Form Submitted Time' : feedback.timestamp
            }
            # Create the email body with an HTML table
            email_body = """\
            <html>
                <body>
                    <h3>Feedback Submission</h3>
                    <table style="border-collapse: collapse; width: 100%;">
                        <tr>
                            <th style="border: 1px solid #dddddd; text-align: left; padding: 8px;">Field</th>
                            <th style="border: 1px solid #dddddd; text-align: left; padding: 8px;">Response</th>
                        </tr>
            """

            for key, value in form_data.items():
                email_body += f"""
                        <tr>
                            <td style="border: 1px solid #dddddd; padding: 8px;">{key}</td>
                            <td style="border: 1px solid #dddddd; padding: 8px;">{value}</td>
                        </tr>
                """

            email_body += """
                    </table>
                </body>
            </html>
            """

            msg = Message("New Feedback Submitted",
                        recipients=["test@example.com"])
            msg.html = email_body
            try:
                mail.send(msg)
            except Exception as e:
                flash(f'Error sending feedback: {str(e)}', 'danger')
                return redirect(url_for('feedback_form'))

            #Redirect to the thank you page
            response = make_response(redirect(url_for('thank_you')))
            response.headers['Cache-Control']='no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
            response.headers['Pragma']='no-cache'
            return response
        
        return render_template('feedback_form.html', form=form)

#Route for thank you page
@app.route('/thank_you')
def thank_you():
        return render_template('thank_you.html')

# Route for admin login
@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        # Dummy credentials for admin login
        if username == 'admin' and password == 'admin':
            session['logged_in'] = True
            session.permanent=True
            return redirect(url_for('admin_panel'))
        else:
            flash('Invalid username or password')
    return render_template('login.html', form=form)

# Admin panel route (requires login)
@app.route('/admin')
def admin_panel():
    if 'logged_in' not in session or not session['logged_in']:
        flash('Please log in to access the admin panel.')
        return redirect(url_for('login'))
    feedbacks = Feedback.query.order_by(Feedback.timestamp.desc()).all()
    return render_template('admin.html', feedbacks=feedbacks)

#export excel route
@app.route('/admin/export', methods=['GET'])
def export_feedback():
    feedbacks = Feedback.query.order_by(Feedback.timestamp.desc()).all()  # Adjust based on your database model

    data = [{
        'ID': index+1,
        'Customer Name': feedback.customer_name,
        'Project': feedback.project,
        'Product': feedback.product,
        'AIPL Job Reference': feedback.aipl_job_reference,
        'Average Rating': feedback.average_rating(),
        'Pre-Order Service': feedback.pre_order_service,
        'Post-Order Handling': feedback.post_order_handling,
        'On-Time Delivery': feedback.on_time_delivery,
        'Quality Sustenance': feedback.quality_sustenance,
        'Overall Response': feedback.overall_response,
        'Final Documentation': feedback.final_documentation,
        'Appraisal': feedback.appraisal,
        'Standard Compliance': feedback.compliance,
        'Feedback submitted Time' : feedback.timestamp
    } for index,feedback in enumerate(feedbacks)]

    df = pd.DataFrame(data)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create a DataFrame and write to the Excel sheet
        df = pd.DataFrame(data)
        df.to_excel(writer, index=False, sheet_name='Feedback')

    # Load the workbook to apply formatting
    output.seek(0)
    workbook = load_workbook(output)
    sheet = workbook.active

    # Create a title row and apply formatting
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Set the header titles and format them
    for col_num, header in enumerate(df.columns, start=1):
        cell = sheet.cell(row=2, column=col_num)  # Start writing headers in row 2
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        # Set the header value
        cell.value = header

    # Set the title for the sheet
    title_cell = sheet.cell(row=1, column=1, value="Customer Feedback Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # Adjust column widths and handle merged cells safely
    for col_num in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = sheet.cell(row=2, column=col_num).column_letter  # Use .column_letter to get the column letter
        for row in range(3, len(data) + 3):  # Start from row 3 for data
            cell_value = sheet.cell(row=row, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Wrap text for all cells
    for row in sheet.iter_rows(min_row=3, max_col=len(df.columns), max_row=len(data) + 2):  # Start from row 3 for data
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Save the workbook again after formatting
    workbook.save(output)

    # Move the pointer to the beginning of the stream
    output.seek(0)

    # Send the file as a response with download_name
    return send_file(output, download_name='feedbacks.xlsx', as_attachment=True)

# Route for logging out
@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
